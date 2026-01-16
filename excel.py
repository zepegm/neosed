import win32com.client
import openpyxl
import os

def rgb_to_hex(rgb):
    bgr = (int(rgb[2]), int(rgb[1]), int(rgb[0]))
    strValue = '%02x%02x%02x' % bgr
    iValue = int(strValue, 16)
    return iValue

class open_xls:
    def __init__(self, filepath):
        self.wb = openpyxl.load_workbook(filepath)
        self.sheet_obj = self.wb.active

    def getCell(self, r, c):
        cell_obj = self.sheet_obj.cell(row=r, column=c)
        return cell_obj.value
    
    def getTotalRows(self):
        return self.sheet_obj.max_row
    
    def getTotalColumns(self):
        return self.sheet_obj.max_column


class xls:
    def __init__(self, planilha=None):
        self.app = win32com.client.Dispatch("Excel.Application")
        self.app.Visible = True  # Torna o Excel visível para depuração

        if planilha is None:
            if self.app.Workbooks.Count == 0:
                raise Exception("Nenhuma planilha está aberta no Excel.")
            self.wb = self.app.ActiveWorkbook
        else:
            caminho_absoluto = os.path.abspath(planilha)

            if not os.path.exists(caminho_absoluto):
                raise FileNotFoundError(f"Arquivo não encontrado: {caminho_absoluto}")

            self.wb = self.app.Workbooks.Open(caminho_absoluto)

        if self.wb is None:
            raise Exception("Falha ao conectar à planilha. Verifique se ela está aberta ou acessível.")

    def getCountA(self, endereco, name=1):
        return self.app.WorksheetFunction.CountA(self.wb.Worksheets(name).Range(endereco))

    def getValCell(self, endereco, name=1):
        return self.wb.Worksheets(name).Range(endereco).value

    def getValCellNumbes(self, name, endereco, row, col):
        return self.wb.Worksheets(name).Range(endereco).Cells(row, col).value

    def setValCell(self, endereco, val, name=1):
        self.wb.Worksheets(name).Range(endereco).value = val

    def setValCellNumbers(self, endereco, val, row, col, name=1):
        self.wb.Worksheets(name).Range(endereco).Cells(row, col).value = val

    def getActiveRow(self):
        return self.app.ActiveCell.Row

    def setRowHeight(self, row, height):
        self.wb.Rows(f"{row}:{row}").RowHeight = height

    def insertPhoto(self, name, caminho, endereco):
        print(f"Inserindo imagem: {caminho}")
        photo = self.wb.Worksheets(name).Pictures().Insert(caminho)
        photo.Top = self.wb.Worksheets(name).Range(endereco).Top
        photo.Left = self.wb.Worksheets(name).Range(endereco).Left + 3
        photo.Height = 516.49
        photo.Width = 569.5

    def setColorBackground(self, name, endereco, rgb):
        self.wb.Worksheets(name).Range(endereco).Interior.Color = rgb_to_hex(rgb)

    def close(self):
        self.wb.Close(SaveChanges=0)
        #self.app.Quit()